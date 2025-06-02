import os
import logging
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.orm import DeclarativeBase
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# Configure logging for debugging
logging.basicConfig(level=logging.DEBUG)

class Base(DeclarativeBase):
    pass

db = SQLAlchemy(model_class=Base)

app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET", "dev-secret-key-change-in-production")

# Configure the database
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL")
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_recycle": 300,
    "pool_pre_ping": True,
}

# Initialize the app with the extension
db.init_app(app)

def init_db():
    """Initialize database tables"""
    with app.app_context():
        # Import models to ensure they're registered
        import models
        # Create all tables
        db.create_all()

@app.route('/')
def index():
    """Dashboard with low stock alerts and recent activity"""
    conn = get_db_connection()
    
    # Get low stock products
    low_stock_products = conn.execute('''
        SELECT code, name, quantity, min_stock, unit, location 
        FROM products 
        WHERE quantity <= min_stock
        ORDER BY quantity ASC
    ''').fetchall()
    
    # Get total products count
    total_products = conn.execute('SELECT COUNT(*) as count FROM products').fetchone()['count']
    
    # Get recent consumption bills
    recent_bills = conn.execute('''
        SELECT id, bill_date, employee_name, is_finished 
        FROM consumption_bills 
        ORDER BY bill_date DESC 
        LIMIT 5
    ''').fetchall()
    
    conn.close()
    
    # Get recent receptions
    reception_conn = get_reception_db_connection()
    recent_receptions = reception_conn.execute('''
        SELECT id, reception_date, supplier, is_finished 
        FROM reception_sheets 
        ORDER BY reception_date DESC 
        LIMIT 5
    ''').fetchall()
    reception_conn.close()
    
    return render_template('index.html', 
                         low_stock_products=low_stock_products,
                         total_products=total_products,
                         recent_bills=recent_bills,
                         recent_receptions=recent_receptions)

@app.route('/products')
def products():
    """Display all products with search functionality"""
    search_query = request.args.get('search', '')
    
    conn = get_db_connection()
    
    if search_query:
        products = conn.execute('''
            SELECT * FROM products 
            WHERE code LIKE ? OR name LIKE ? OR location LIKE ?
            ORDER BY name
        ''', (f'%{search_query}%', f'%{search_query}%', f'%{search_query}%')).fetchall()
    else:
        products = conn.execute('SELECT * FROM products ORDER BY name').fetchall()
    
    conn.close()
    
    return render_template('products.html', products=products, search_query=search_query)

@app.route('/products/add', methods=['GET', 'POST'])
def add_product():
    """Add new product"""
    if request.method == 'POST':
        code = request.form['code'].strip()
        name = request.form['name'].strip()
        unit = request.form['unit'].strip()
        quantity = float(request.form['quantity'])
        location = request.form['location'].strip()
        min_stock = float(request.form['min_stock'])
        
        conn = get_db_connection()
        try:
            conn.execute('''
                INSERT INTO products (code, name, unit, quantity, location, min_stock)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (code, name, unit, quantity, location, min_stock))
            conn.commit()
            flash('Produsul a fost adăugat cu succes!', 'success')
            return redirect(url_for('products'))
        except sqlite3.IntegrityError:
            flash('Codul produsului există deja!', 'error')
        finally:
            conn.close()
    
    return render_template('products.html', action='add')

@app.route('/products/edit/<int:product_id>', methods=['GET', 'POST'])
def edit_product(product_id):
    """Edit existing product"""
    conn = get_db_connection()
    
    if request.method == 'POST':
        code = request.form['code'].strip()
        name = request.form['name'].strip()
        unit = request.form['unit'].strip()
        quantity = float(request.form['quantity'])
        location = request.form['location'].strip()
        min_stock = float(request.form['min_stock'])
        
        try:
            conn.execute('''
                UPDATE products 
                SET code=?, name=?, unit=?, quantity=?, location=?, min_stock=?
                WHERE id=?
            ''', (code, name, unit, quantity, location, min_stock, product_id))
            conn.commit()
            flash('Produsul a fost actualizat cu succes!', 'success')
            return redirect(url_for('products'))
        except sqlite3.IntegrityError:
            flash('Codul produsului există deja!', 'error')
        finally:
            conn.close()
    
    product = conn.execute('SELECT * FROM products WHERE id = ?', (product_id,)).fetchone()
    conn.close()
    
    if not product:
        flash('Produsul nu a fost găsit!', 'error')
        return redirect(url_for('products'))
    
    return render_template('products.html', action='edit', product=product)

@app.route('/products/delete/<int:product_id>')
def delete_product(product_id):
    """Delete product"""
    conn = get_db_connection()
    conn.execute('DELETE FROM products WHERE id = ?', (product_id,))
    conn.commit()
    conn.close()
    
    flash('Produsul a fost șters cu succes!', 'success')
    return redirect(url_for('products'))

@app.route('/consumption_bills')
def consumption_bills():
    """Display all consumption bills"""
    conn = get_db_connection()
    bills = conn.execute('''
        SELECT id, bill_date, employee_name, employee_signature, is_finished
        FROM consumption_bills 
        ORDER BY bill_date DESC
    ''').fetchall()
    conn.close()
    
    return render_template('consumption_bills.html', bills=bills)

@app.route('/consumption_bills/create')
def create_consumption_bill():
    """Create new consumption bill"""
    # Load draft if exists
    draft_data = load_draft_bill()
    
    conn = get_db_connection()
    products = conn.execute('SELECT * FROM products ORDER BY name').fetchall()
    conn.close()
    
    return render_template('bill_create.html', products=products, draft_data=draft_data)

@app.route('/consumption_bills/add_item', methods=['POST'])
def add_bill_item():
    """Add item to current bill (AJAX endpoint)"""
    product_code = request.form['product_code']
    quantity = float(request.form['quantity'])
    
    conn = get_db_connection()
    product = conn.execute('SELECT * FROM products WHERE code = ?', (product_code,)).fetchone()
    conn.close()
    
    if not product:
        return jsonify({'error': 'Produsul nu a fost găsit'}), 400
    
    if quantity > product['quantity']:
        return jsonify({'error': 'Cantitatea solicitată depășește stocul disponibil'}), 400
    
    # Add to session
    if 'bill_items' not in session:
        session['bill_items'] = []
    
    item_number = len(session['bill_items']) + 1
    item = {
        'item_number': item_number,
        'code': product['code'],
        'name': product['name'],
        'unit': product['unit'],
        'quantity': quantity,
        'location': product['location']
    }
    
    session['bill_items'].append(item)
    session.modified = True
    
    return jsonify({'success': True, 'item': item})

@app.route('/consumption_bills/remove_item/<int:item_index>')
def remove_bill_item(item_index):
    """Remove item from current bill"""
    if 'bill_items' in session and 0 <= item_index < len(session['bill_items']):
        session['bill_items'].pop(item_index)
        # Renumber items
        for i, item in enumerate(session['bill_items']):
            item['item_number'] = i + 1
        session.modified = True
        flash('Articolul a fost eliminat!', 'success')
    
    return redirect(url_for('create_consumption_bill'))

@app.route('/consumption_bills/save_draft', methods=['POST'])
def save_bill_draft():
    """Save current bill as draft"""
    employee_name = request.form.get('employee_name', '')
    employee_signature = request.form.get('employee_signature', '')
    
    conn = get_db_connection()
    
    # Clear existing draft
    conn.execute('DELETE FROM draft_bill_items')
    conn.execute('DELETE FROM draft_bills')
    
    # Save new draft
    cursor = conn.execute('''
        INSERT INTO draft_bills (employee_name, employee_signature, last_updated)
        VALUES (?, ?, ?)
    ''', (employee_name, employee_signature, datetime.now().isoformat()))
    
    draft_id = cursor.lastrowid
    
    # Save draft items
    if 'bill_items' in session:
        for item in session['bill_items']:
            conn.execute('''
                INSERT INTO draft_bill_items 
                (draft_id, item_number, product_code, product_name, unit, quantity, location)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (draft_id, item['item_number'], item['code'], item['name'], 
                  item['unit'], item['quantity'], item['location']))
    
    conn.commit()
    conn.close()
    
    flash('Bonul a fost salvat ca ciornă!', 'success')
    return redirect(url_for('create_consumption_bill'))

@app.route('/consumption_bills/finalize', methods=['POST'])
def finalize_consumption_bill():
    """Finalize consumption bill"""
    employee_name = request.form['employee_name'].strip()
    employee_signature = request.form['employee_signature'].strip()
    
    if not employee_name:
        flash('Numele angajatului este obligatoriu!', 'error')
        return redirect(url_for('create_consumption_bill'))
    
    if 'bill_items' not in session or not session['bill_items']:
        flash('Nu există articole în bon!', 'error')
        return redirect(url_for('create_consumption_bill'))
    
    conn = get_db_connection()
    
    try:
        # Create bill
        cursor = conn.execute('''
            INSERT INTO consumption_bills (bill_date, employee_name, employee_signature, is_finished)
            VALUES (?, ?, ?, 1)
        ''', (datetime.now().isoformat(), employee_name, employee_signature))
        
        bill_id = cursor.lastrowid
        
        # Add bill items and update stock
        for item in session['bill_items']:
            # Add item to bill
            conn.execute('''
                INSERT INTO bill_items 
                (bill_id, item_number, product_code, product_name, unit, quantity, location)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (bill_id, item['item_number'], item['code'], item['name'],
                  item['unit'], item['quantity'], item['location']))
            
            # Update product stock
            conn.execute('''
                UPDATE products 
                SET quantity = quantity - ?
                WHERE code = ?
            ''', (item['quantity'], item['code']))
        
        # Clear draft
        conn.execute('DELETE FROM draft_bill_items')
        conn.execute('DELETE FROM draft_bills')
        
        conn.commit()
        
        # Clear session
        session.pop('bill_items', None)
        
        flash('Bonul de consum a fost finalizat cu succes!', 'success')
        return redirect(url_for('consumption_bills'))
        
    except Exception as e:
        conn.rollback()
        flash(f'Eroare la finalizarea bonului: {str(e)}', 'error')
        return redirect(url_for('create_consumption_bill'))
    finally:
        conn.close()

@app.route('/consumption_bills/view/<int:bill_id>')
def view_consumption_bill(bill_id):
    """View consumption bill details"""
    conn = get_db_connection()
    
    bill = conn.execute('''
        SELECT * FROM consumption_bills WHERE id = ?
    ''', (bill_id,)).fetchone()
    
    if not bill:
        flash('Bonul nu a fost găsit!', 'error')
        return redirect(url_for('consumption_bills'))
    
    items = conn.execute('''
        SELECT * FROM bill_items WHERE bill_id = ? ORDER BY item_number
    ''', (bill_id,)).fetchall()
    
    conn.close()
    
    return render_template('bill_create.html', bill=bill, items=items, view_mode=True)

@app.route('/consumption_bills/export/<int:bill_id>')
def export_consumption_bill(bill_id):
    """Export consumption bill to Excel"""
    conn = get_db_connection()
    
    bill = conn.execute('SELECT * FROM consumption_bills WHERE id = ?', (bill_id,)).fetchone()
    items = conn.execute('SELECT * FROM bill_items WHERE bill_id = ? ORDER BY item_number', (bill_id,)).fetchall()
    
    conn.close()
    
    if not bill:
        flash('Bonul nu a fost găsit!', 'error')
        return redirect(url_for('consumption_bills'))
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Bon Consum {bill_id}"
    
    # Header
    ws['A1'] = 'BON DE CONSUM'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Bill info
    ws['A3'] = f"Data: {bill['bill_date']}"
    ws['A4'] = f"Angajat: {bill['employee_name']}"
    ws['A5'] = f"Semnătura: {bill['employee_signature']}"
    
    # Table header
    headers = ['Nr.', 'Cod Produs', 'Denumire', 'U.M.', 'Cantitate', 'Locație']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
    
    # Table data
    for row, item in enumerate(items, 8):
        ws.cell(row=row, column=1, value=item['item_number'])
        ws.cell(row=row, column=2, value=item['product_code'])
        ws.cell(row=row, column=3, value=item['product_name'])
        ws.cell(row=row, column=4, value=item['unit'])
        ws.cell(row=row, column=5, value=item['quantity'])
        ws.cell(row=row, column=6, value=item['location'])
        
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(
                top=Side(style='thin'),
                bottom=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )
    
    # Adjust column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'bon_consum_{bill_id}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/reception')
def reception():
    """Display all reception sheets"""
    reception_conn = get_reception_db_connection()
    receptions = reception_conn.execute('''
        SELECT id, reception_date, supplier, document_number, notes, is_finished
        FROM reception_sheets 
        ORDER BY reception_date DESC
    ''').fetchall()
    reception_conn.close()
    
    return render_template('reception.html', receptions=receptions)

@app.route('/reception/view/<int:reception_id>')
def view_reception(reception_id):
    """View reception details"""
    reception_conn = get_reception_db_connection()
    
    reception = reception_conn.execute('''
        SELECT * FROM reception_sheets WHERE id = ?
    ''', (reception_id,)).fetchone()
    
    if not reception:
        reception_conn.close()
        return jsonify({'error': 'Recepția nu a fost găsită'}), 404
    
    items = reception_conn.execute('''
        SELECT * FROM reception_items WHERE reception_id = ? ORDER BY item_number
    ''', (reception_id,)).fetchall()
    
    reception_conn.close()
    
    # Return JSON for AJAX request
    return jsonify({
        'reception': {
            'id': reception['id'],
            'date': reception['reception_date'],
            'supplier': reception['supplier'],
            'document_number': reception['document_number'],
            'notes': reception['notes'],
            'is_finished': reception['is_finished']
        },
        'items': [dict(item) for item in items]
    })

@app.route('/reception/export/<int:reception_id>')
def export_reception(reception_id):
    """Export reception to Excel"""
    reception_conn = get_reception_db_connection()
    
    reception = reception_conn.execute('SELECT * FROM reception_sheets WHERE id = ?', (reception_id,)).fetchone()
    items = reception_conn.execute('SELECT * FROM reception_items WHERE reception_id = ? ORDER BY item_number', (reception_id,)).fetchall()
    
    reception_conn.close()
    
    if not reception:
        flash('Recepția nu a fost găsită!', 'error')
        return redirect(url_for('reception'))
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Receptie {reception_id}"
    
    # Header
    ws['A1'] = 'FIȘĂ DE RECEPȚIE'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Reception info
    ws['A3'] = f"Data: {reception['reception_date']}"
    ws['A4'] = f"Furnizor: {reception['supplier']}"
    ws['A5'] = f"Nr. Document: {reception['document_number'] or '-'}"
    ws['A6'] = f"Observații: {reception['notes'] or '-'}"
    
    # Table header
    headers = ['Nr.', 'Cod Produs', 'Denumire', 'U.M.', 'Cantitate', 'Locație', 'Data Intrare']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
    
    # Table data
    for row, item in enumerate(items, 9):
        ws.cell(row=row, column=1, value=item['item_number'])
        ws.cell(row=row, column=2, value=item['product_code'])
        ws.cell(row=row, column=3, value=item['product_name'])
        ws.cell(row=row, column=4, value=item['unit'])
        ws.cell(row=row, column=5, value=item['quantity'])
        ws.cell(row=row, column=6, value=item['location'] or '-')
        ws.cell(row=row, column=7, value=item['entry_date'][:10] if item['entry_date'] else '-')
        
        # Add borders
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = Border(
                top=Side(style='thin'),
                bottom=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )
    
    # Auto-adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'receptie_{reception_id}_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/reception/save_draft', methods=['POST'])
def save_reception_draft():
    """Save current reception as draft"""
    supplier = request.form.get('supplier', '')
    document_number = request.form.get('document_number', '')
    notes = request.form.get('notes', '')
    
    reception_conn = get_reception_db_connection()
    
    # Clear existing draft
    reception_conn.execute('DELETE FROM draft_reception_items')
    reception_conn.execute('DELETE FROM draft_receptions')
    
    # Save new draft
    cursor = reception_conn.execute('''
        INSERT INTO draft_receptions (supplier, document_number, notes, last_updated)
        VALUES (?, ?, ?, ?)
    ''', (supplier, document_number, notes, datetime.now().isoformat()))
    
    draft_id = cursor.lastrowid
    
    # Save draft items
    if 'reception_items' in session:
        for item in session['reception_items']:
            reception_conn.execute('''
                INSERT INTO draft_reception_items 
                (draft_id, item_number, product_code, product_name, unit, quantity, location)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (draft_id, item['item_number'], item['code'], item['name'], 
                  item['unit'], item['quantity'], item['location']))
    
    reception_conn.commit()
    reception_conn.close()
    
    flash('Recepția a fost salvată ca ciornă!', 'success')
    return redirect(url_for('create_reception'))

@app.route('/reception/create')
def create_reception():
    """Create new reception sheet"""
    # Load draft if exists
    draft_data = load_draft_reception()
    
    conn = get_db_connection()
    products = conn.execute('SELECT * FROM products ORDER BY name').fetchall()
    conn.close()
    
    return render_template('reception_create.html', products=products, draft_data=draft_data)

@app.route('/reception/add_item', methods=['POST'])
def add_reception_item():
    """Add item to current reception"""
    product_code = request.form['product_code']
    quantity = float(request.form['quantity'])
    
    conn = get_db_connection()
    product = conn.execute('SELECT * FROM products WHERE code = ?', (product_code,)).fetchone()
    conn.close()
    
    if not product:
        return jsonify({'error': 'Produsul nu a fost găsit'}), 400
    
    # Add to session
    if 'reception_items' not in session:
        session['reception_items'] = []
    
    item_number = len(session['reception_items']) + 1
    item = {
        'item_number': item_number,
        'code': product['code'],
        'name': product['name'],
        'unit': product['unit'],
        'quantity': quantity,
        'location': product['location']
    }
    
    session['reception_items'].append(item)
    session.modified = True
    
    return jsonify({'success': True, 'item': item})

@app.route('/reception/remove_item/<int:item_index>')
def remove_reception_item(item_index):
    """Remove item from current reception"""
    if 'reception_items' in session and 0 <= item_index < len(session['reception_items']):
        session['reception_items'].pop(item_index)
        # Renumber items
        for i, item in enumerate(session['reception_items']):
            item['item_number'] = i + 1
        session.modified = True
        flash('Articolul a fost eliminat!', 'success')
    
    return redirect(url_for('create_reception'))

@app.route('/reception/finalize', methods=['POST'])
def finalize_reception():
    """Finalize reception sheet"""
    supplier = request.form['supplier'].strip()
    document_number = request.form['document_number'].strip()
    notes = request.form['notes'].strip()
    
    if not supplier:
        flash('Furnizorul este obligatoriu!', 'error')
        return redirect(url_for('create_reception'))
    
    if 'reception_items' not in session or not session['reception_items']:
        flash('Nu există articole în recepție!', 'error')
        return redirect(url_for('create_reception'))
    
    reception_conn = get_reception_db_connection()
    conn = get_db_connection()
    
    try:
        # Create reception
        cursor = reception_conn.execute('''
            INSERT INTO reception_sheets (reception_date, supplier, document_number, notes, is_finished)
            VALUES (?, ?, ?, ?, 1)
        ''', (datetime.now().isoformat(), supplier, document_number, notes))
        
        reception_id = cursor.lastrowid
        
        # Add reception items and update stock
        for item in session['reception_items']:
            # Add item to reception
            reception_conn.execute('''
                INSERT INTO reception_items 
                (reception_id, item_number, product_code, product_name, unit, quantity, location, entry_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (reception_id, item['item_number'], item['code'], item['name'],
                  item['unit'], item['quantity'], item['location'], datetime.now().isoformat()))
            
            # Update product stock
            conn.execute('''
                UPDATE products 
                SET quantity = quantity + ?
                WHERE code = ?
            ''', (item['quantity'], item['code']))
        
        # Clear draft
        reception_conn.execute('DELETE FROM draft_reception_items')
        reception_conn.execute('DELETE FROM draft_receptions')
        
        reception_conn.commit()
        conn.commit()
        
        # Clear session
        session.pop('reception_items', None)
        
        flash('Fișa de recepție a fost finalizată cu succes!', 'success')
        return redirect(url_for('reception'))
        
    except Exception as e:
        reception_conn.rollback()
        conn.rollback()
        flash(f'Eroare la finalizarea recepției: {str(e)}', 'error')
        return redirect(url_for('create_reception'))
    finally:
        reception_conn.close()
        conn.close()

def load_draft_bill():
    """Load draft bill data"""
    conn = get_db_connection()
    draft = conn.execute('SELECT id FROM draft_bills ORDER BY last_updated DESC LIMIT 1').fetchone()
    
    if not draft:
        conn.close()
        return None
    
    draft_id = draft['id']
    
    # Get draft info
    draft_info = conn.execute('SELECT * FROM draft_bills WHERE id = ?', (draft_id,)).fetchone()
    
    # Get draft items
    items = conn.execute('SELECT * FROM draft_bill_items WHERE draft_id = ?', (draft_id,)).fetchall()
    
    conn.close()
    
    # Convert to session format
    session['bill_items'] = []
    for item in items:
        session['bill_items'].append({
            'item_number': item['item_number'],
            'code': item['product_code'],
            'name': item['product_name'],
            'unit': item['unit'],
            'quantity': item['quantity'],
            'location': item['location']
        })
    
    return {
        'employee_name': draft_info['employee_name'],
        'employee_signature': draft_info['employee_signature']
    }

def load_draft_reception():
    """Load draft reception data"""
    reception_conn = get_reception_db_connection()
    draft = reception_conn.execute('SELECT id FROM draft_receptions ORDER BY last_updated DESC LIMIT 1').fetchone()
    
    if not draft:
        reception_conn.close()
        return None
    
    draft_id = draft['id']
    
    # Get draft info
    draft_info = reception_conn.execute('SELECT * FROM draft_receptions WHERE id = ?', (draft_id,)).fetchone()
    
    # Get draft items
    items = reception_conn.execute('SELECT * FROM draft_reception_items WHERE draft_id = ?', (draft_id,)).fetchall()
    
    reception_conn.close()
    
    # Convert to session format
    session['reception_items'] = []
    for item in items:
        session['reception_items'].append({
            'item_number': item['item_number'],
            'code': item['product_code'],
            'name': item['product_name'],
            'unit': item['unit'],
            'quantity': item['quantity'],
            'location': item['location']
        })
    
    return {
        'supplier': draft_info['supplier'],
        'document_number': draft_info['document_number'],
        'notes': draft_info['notes']
    }

# Initialize database on startup
init_db()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
